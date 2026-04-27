import openpyxl
from decimal import Decimal, InvalidOperation
from django.db import transaction
from django.views import View
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.http import JsonResponse
from core import models

IMPORT_HEADER_ROW_INDEX = 4  # 0-based → row 5 in Excel


def _imp_dec(val):
    try:
        if val in (None, "", "-", "#DIV/0!"):
            return None
        return Decimal(str(val))
    except InvalidOperation:
        return None


def _imp_int(val):
    try:
        if val in (None, "", "-"):
            return None
        return int(float(str(val)))
    except (ValueError, TypeError):
        return None


def _imp_str(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s and s.lower() != "nan" else None


def _imp_bool(val):
    s = _imp_str(val)
    if s is None:
        return None
    return s.upper() not in ("NO", "0", "FALSE", "N")


def _imp_ok(val):
    s = _imp_str(val)
    return s if s in ("ok", "not ok") else None


def _imp_month(raw):
    if raw is None:
        raise ValueError("Month value is missing.")
    d = raw.date() if hasattr(raw, "date") else raw
    first = d.replace(day=1)
    obj, _ = models.Month.objects.get_or_create(
        month_date=first,
        defaults={"label": first.strftime("%b %Y")},
    )
    return obj


@method_decorator(csrf_exempt, name="dispatch")
class ExcelImportView(View):
    """
    POST /app/api/import/excel/
    Accepts a multipart .xlsx upload and upserts Campaign, Metric, Comparison.
    Excel structure: headers on row 5, data from row 6 (Register2026 sheet).
    """

    def post(self, request):
        file = request.FILES.get("file")
        if not file:
            return JsonResponse({"error": "No file provided."}, status=400)
        if not (file.name.endswith(".xlsx") or file.name.endswith(".xls")):
            return JsonResponse({"error": "File must be .xlsx or .xls"}, status=400)

        try:
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb["Register2026"] if "Register2026" in wb.sheetnames else wb.active
        except Exception as exc:
            return JsonResponse({"error": f"Failed to read file: {exc}"}, status=400)

        all_rows = list(ws.iter_rows(values_only=True))
        if len(all_rows) <= IMPORT_HEADER_ROW_INDEX:
            return JsonResponse({"error": "File format not recognised."}, status=400)

        raw_headers = all_rows[IMPORT_HEADER_ROW_INDEX]
        headers = [
            str(h).strip().replace("\n", " ") if h is not None else ""
            for h in raw_headers
        ]
        data_rows = all_rows[IMPORT_HEADER_ROW_INDEX + 1 :]
        col = {h: i for i, h in enumerate(headers)}

        def get(row, name, fallback=None):
            idx = col.get(name)
            if idx is None:
                return fallback
            v = row[idx] if idx < len(row) else None
            return v if v is not None else fallback

        created, updated, errors = 0, 0, []

        with transaction.atomic():
            for row_num, row in enumerate(data_rows, start=IMPORT_HEADER_ROW_INDEX + 2):
                if all(v is None for v in row):
                    continue
                try:
                    month = _imp_month(get(row, "Month"))

                    country_code = _imp_str(get(row, "Country")) or "XX"
                    country, _ = models.Country.objects.get_or_create(
                        iso_code=country_code, defaults={"name": country_code}
                    )

                    client_code = _imp_str(get(row, "Client")) or "UNKNOWN"
                    client, _ = models.Client.objects.get_or_create(
                        code=client_code, defaults={"name": client_code}
                    )

                    cat_name = _imp_str(get(row, "Categoria"))
                    category = (
                        models.Category.objects.get_or_create(name=cat_name)[0]
                        if cat_name
                        else None
                    )

                    creative_name = _imp_str(get(row, "Creativity"))
                    creative = (
                        models.Creative.objects.get_or_create(name=creative_name)[0]
                        if creative_name
                        else None
                    )

                    platform_name = _imp_str(get(row, "Platform"))
                    platform = (
                        models.Platform.objects.get_or_create(name=platform_name)[0]
                        if platform_name
                        else None
                    )

                    plat_owner_name = _imp_str(get(row, "Plat. owner"))
                    plat_owner = (
                        models.PlatOwner.objects.get_or_create(name=plat_owner_name)[0]
                        if plat_owner_name
                        else None
                    )

                    detail_name = _imp_str(get(row, "Detalhe"))
                    detail = (
                        models.DetailType.objects.get_or_create(name=detail_name)[0]
                        if detail_name
                        else None
                    )

                    team, _ = models.Team.objects.get_or_create(
                        ts_manager=_imp_str(get(row, "TS Manager")),
                        ts_head=_imp_str(get(row, "TS Head")),
                        client_manager=_imp_str(get(row, "Client Manager")),
                        client_head=_imp_str(get(row, "Client Head")),
                        fe=_imp_str(get(row, "FE")),
                        be=_imp_str(get(row, "BE")),
                        creative=_imp_str(get(row, "Criativo")),
                        qa=_imp_str(get(row, "QA")),
                        pm=_imp_str(get(row, "PM")),
                    )

                    raw_invoice = get(row, "Invoice google")
                    invoice_str = (
                        str(int(float(str(raw_invoice))))
                        if raw_invoice not in (None, "")
                        else None
                    )
                    if invoice_str and len(invoice_str) > 20:
                        invoice_str = invoice_str[:20]

                    campaign = models.Campaign.objects.filter(
                        month=month,
                        client=client,
                        country=country,
                        creative=creative,
                        category=category,
                    ).first()

                    campaign_defaults = {
                        "platform": platform,
                        "plat_owner": plat_owner,
                        "team": team,
                        "detail_type": detail,
                        "client_camp": _imp_bool(get(row, "Client Camp.")),
                        "invoice_google": invoice_str,
                    }

                    if campaign:
                        for attr, val in campaign_defaults.items():
                            setattr(campaign, attr, val)
                        campaign.save()
                        updated += 1
                    else:
                        campaign = models.Campaign.objects.create(
                            month=month,
                            client=client,
                            country=country,
                            creative=creative,
                            category=category,
                            **campaign_defaults,
                        )
                        created += 1

                    models.Metric.objects.update_or_create(
                        campaign=campaign,
                        defaults={
                            "cpa": _imp_dec(get(row, "CPA")),
                            "er_cpa": _imp_dec(get(row, "ER cpa")),
                            "er_cost": _imp_dec(get(row, "ER cost")),
                            "cost_ts": _imp_dec(get(row, "Cost TS")),
                            "conv": _imp_int(get(row, "Conv")),
                            "cost_eur": _imp_dec(get(row, "Cost €")),
                            "revenue_eur": _imp_dec(get(row, "Revenue €")),
                            "margin_eur": _imp_dec(get(row, "Margin €")),
                            "roi": _imp_dec(get(row, "ROI")),
                        },
                    )

                    models.Comparison.objects.update_or_create(
                        campaign=campaign,
                        defaults={
                            "client_conv": _imp_int(get(row, "conv cliente")),
                            "client_rev": _imp_dec(get(row, "rev cliente")),
                            "var_conv": _imp_dec(get(row, "var conv")),
                            "var_rev": _imp_dec(get(row, "var rev")),
                            "var_conv_pct": _imp_dec(get(row, "var conv %")),
                            "var_rev_pct": _imp_dec(get(row, "var rev %")),
                            "ok_conv": _imp_ok(get(row, "ok/not ok conv")),
                            "ok_rev": _imp_ok(get(row, "ok/not ok rev")),
                        },
                    )

                except Exception as exc:
                    errors.append({"row": row_num, "error": str(exc)})

        return JsonResponse(
            {
                "created": created,
                "updated": updated,
                "errors": errors,
                "total_processed": created + updated + len(errors),
            }
        )
