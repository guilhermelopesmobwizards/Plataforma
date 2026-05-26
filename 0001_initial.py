"""
core/export/excel_export.py

Generates the monthly closing Excel file matching the MEC Register2026 format exactly.
Two sheets: Register2026 (full data) + Pivots_2026 (monthly summary).
"""

from __future__ import annotations

import io
from decimal import Decimal
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core import models


# ── Colour palette ────────────────────────────────────────────────────────────
_PURPLE     = "9751CB"   # Cost TS, Conv header bg
_PURPLE_TXT = "F2F2F2"  # text on purple
_PEACH      = "FBE4D5"  # conv cliente, rev cliente header bg
_ORANGE     = "F1995D"  # var cols, ok/not ok header bg
_GREY_SUB   = "D0CECE"  # subtotal row bg
_BLACK      = "000000"
_GREY_DARK  = "424242"

# ── Column indices (A is narrow spacer, data starts at B=2) ──────────────────
COL_MONTH        = 2
COL_CLIENT       = 3
COL_COUNTRY      = 4
COL_CATEGORIA    = 5
COL_CREATIVITY   = 6
COL_PLATFORM     = 7
COL_PLAT_OWNER   = 8
COL_CLIENT_CAMP  = 9
COL_CPA          = 10
COL_ER_CPA       = 11
COL_ER_COST      = 12
COL_COST_TS      = 13  # M
COL_CONV         = 14  # N
COL_COST_EUR     = 15  # O  → =M*L
COL_REV_EUR      = 16  # P  → =N*K*J
COL_MARGIN_EUR   = 17  # Q  → =P-O
COL_ROI          = 18  # R  → =Q/O
COL_TS_MANAGER   = 19
COL_TS_HEAD      = 20
COL_CLIENT_MANAGER = 21
COL_CLIENT_HEAD  = 22
COL_FE           = 23
COL_BE           = 24
COL_CRIATIVO     = 25
COL_QA           = 26
COL_PM           = 27
COL_DETALHE      = 28
COL_CONV_CLIENT  = 29  # AC
COL_REV_CLIENT   = 30  # AD
COL_VAR_CONV     = 31  # AE → =AC-N
COL_VAR_REV      = 32  # AF → =AD-(N*J)
COL_VAR_CONV_PCT = 33  # AG → =AE/AC
COL_VAR_REV_PCT  = 34  # AH → =AF/AD
COL_OK_CONV      = 35  # AI → IF formula
COL_OK_REV       = 36  # AJ → IF formula
COL_INVOICE      = 37  # AK

HEADER_ROW    = 5
DATA_START    = 6
LAST_DATA_ROW = 14595  # matches original SUBTOTAL range

# (header label, column width)
HEADERS = {
    COL_MONTH:           ("Month",            7.7),
    COL_CLIENT:          ("Client",           6.7),
    COL_COUNTRY:         ("Country",          8.2),
    COL_CATEGORIA:       ("Categoria",       14.0),
    COL_CREATIVITY:      ("Creativity",      16.0),
    COL_PLATFORM:        ("Platform",        10.0),
    COL_PLAT_OWNER:      ("Plat. owner",     12.0),
    COL_CLIENT_CAMP:     ("Client Camp.",    12.0),
    COL_CPA:             ("CPA",             10.0),
    COL_ER_CPA:          ("ER cpa",          10.0),
    COL_ER_COST:         ("ER cost",         10.0),
    COL_COST_TS:         ("Cost TS",         14.0),
    COL_CONV:            ("Conv",            10.0),
    COL_COST_EUR:        ("Cost €",          12.0),
    COL_REV_EUR:         ("Revenue €",       12.0),
    COL_MARGIN_EUR:      ("Margin €",        12.0),
    COL_ROI:             ("ROI",             10.0),
    COL_TS_MANAGER:      ("TS Manager",      12.0),
    COL_TS_HEAD:         ("TS Head",         12.0),
    COL_CLIENT_MANAGER:  ("Client Manager",  14.0),
    COL_CLIENT_HEAD:     ("Client Head",     12.0),
    COL_FE:              ("FE",               8.0),
    COL_BE:              ("BE",               8.0),
    COL_CRIATIVO:        ("Criativo",        10.0),
    COL_QA:              ("QA",               8.0),
    COL_PM:              ("PM",               8.0),
    COL_DETALHE:         ("Detalhe",         10.0),
    COL_CONV_CLIENT:     ("conv cliente",    12.0),
    COL_REV_CLIENT:      ("rev cliente",     12.0),
    COL_VAR_CONV:        ("var conv",        10.0),
    COL_VAR_REV:         ("var rev",         10.0),
    COL_VAR_CONV_PCT:    ("var conv %",      12.0),
    COL_VAR_REV_PCT:     ("var rev %",       12.0),
    COL_OK_CONV:         ("ok/not ok\nconv", 14.0),
    COL_OK_REV:          ("ok/not ok\nrev",  13.0),
    COL_INVOICE:         ("Invoice google",  16.0),
}

HEADER_BG = {
    COL_COST_TS:      _PURPLE,
    COL_CONV:         _PURPLE,
    COL_CONV_CLIENT:  _PEACH,
    COL_REV_CLIENT:   _PEACH,
    COL_VAR_CONV:     _ORANGE,
    COL_VAR_REV:      _ORANGE,
    COL_VAR_CONV_PCT: _ORANGE,
    COL_VAR_REV_PCT:  _ORANGE,
    COL_OK_CONV:      _ORANGE,
    COL_OK_REV:       _ORANGE,
}

HEADER_TXT = {
    COL_COST_TS: _PURPLE_TXT,
    COL_CONV:    _PURPLE_TXT,
}

# Number formats matching MEC exactly
NUM_FMT = {
    COL_CPA:         "#,##0.00",
    COL_ER_CPA:      "#,##0.00",
    COL_ER_COST:     "#,##0.00",
    COL_COST_TS:     "#,##0.0",
    COL_CONV:        "#,##0",
    COL_COST_EUR:    "#,##0",
    COL_REV_EUR:     "#,##0",
    COL_MARGIN_EUR:  "#,##0",
    COL_ROI:         "0%",
    COL_CONV_CLIENT: "#,##0.00",
    COL_REV_CLIENT:  "#,##0.00",
    COL_VAR_CONV:    "#,##0",
    COL_VAR_REV:     "#,##0",
    COL_VAR_CONV_PCT: "0.0%",
    COL_VAR_REV_PCT:  "0.0%",
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def _fill(c: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=c)


def _center(wrap=False) -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)


def _L(col: int) -> str:
    return get_column_letter(col)


def _dec(campaign, field: str):
    m = getattr(campaign, "metric", None)
    if m is None:
        return None
    val = getattr(m, field, None)
    return float(val) if isinstance(val, Decimal) else val


def _int(campaign, field: str):
    m = getattr(campaign, "metric", None)
    return getattr(m, field, None) if m else None


def _cmp(campaign, field: str):
    cp = getattr(campaign, "comparison", None)
    if cp is None:
        return None
    val = getattr(cp, field, None)
    return float(val) if isinstance(val, Decimal) else val


def _worker(campaign, field: str):
    w = campaign.worker
    return getattr(w, field, None) if w else None


# ── Sheet 1: Register2026 ─────────────────────────────────────────────────────

def _build_register(wb, campaigns: list) -> None:
    ws = wb.active
    ws.title = "Register2026"

    # Column A: narrow spacer
    ws.column_dimensions["A"].width = 1.7
    for col_idx, (_, width) in HEADERS.items():
        ws.column_dimensions[_L(col_idx)].width = width

    # Rows 1-4: height 12
    for r in range(1, 5):
        ws.row_dimensions[r].height = 12

    # ── Row 2: KPI reference labels ───────────────────────────────────────────
    total_cost_ts = sum((_dec(c, "cost_ts") or 0) for c in campaigns)
    c2 = ws.cell(row=2, column=COL_COST_TS, value=total_cost_ts)
    c2.font = Font(size=9, name="Calibri")
    c2.number_format = "#,##0.00"

    for col, label in [
        (COL_CONV,        "Conv"),
        (COL_COST_EUR,    "Cost €"),
        (COL_REV_EUR,     "Revenue €"),
        (COL_MARGIN_EUR,  "Margin €"),
        (COL_ROI,         "ROI"),
        (COL_CONV_CLIENT, "conv"),
        (COL_REV_CLIENT,  "rev"),
    ]:
        ws.cell(row=2, column=col, value=label).font = Font(bold=True, size=9, name="Calibri")

    # ── Row 3: SUBTOTAL formulas (grey bg) ────────────────────────────────────
    M  = _L(COL_COST_TS);   N  = _L(COL_CONV)
    O  = _L(COL_COST_EUR);  P  = _L(COL_REV_EUR)
    Q  = _L(COL_MARGIN_EUR); AC = _L(COL_CONV_CLIENT)
    AD = _L(COL_REV_CLIENT)

    for col, formula in [
        (COL_COST_TS,    f"=SUBTOTAL(9,{M}{DATA_START}:{M}{LAST_DATA_ROW})"),
        (COL_CONV,       f"=SUBTOTAL(9,{N}{DATA_START}:{N}{LAST_DATA_ROW})"),
        (COL_COST_EUR,   f"=SUBTOTAL(9,{O}{DATA_START}:{O}{LAST_DATA_ROW})"),
        (COL_REV_EUR,    f"=SUBTOTAL(9,{P}{DATA_START}:{P}{LAST_DATA_ROW})"),
        (COL_MARGIN_EUR, f"=SUBTOTAL(9,{Q}{DATA_START}:{Q}{LAST_DATA_ROW})"),
        (COL_ROI,        f"={Q}3/{O}3"),
        (COL_CONV_CLIENT,f"=SUBTOTAL(9,{AC}{DATA_START}:{AC}{LAST_DATA_ROW})"),
        (COL_REV_CLIENT, f"=SUBTOTAL(9,{AD}{DATA_START}:{AD}{LAST_DATA_ROW})"),
    ]:
        cell = ws.cell(row=3, column=col, value=formula)
        cell.fill = _fill(_GREY_SUB)
        cell.font = Font(bold=True, size=9, name="Calibri")
        cell.alignment = _center()
        fmt = NUM_FMT.get(col, "General")
        cell.number_format = fmt

    # ── Row 4: delta (reference − actual) ────────────────────────────────────
    ws.cell(row=4, column=COL_COST_TS, value=f"=+{M}2-{M}3").font = Font(size=9, name="Calibri")

    # ── Row 5: column headers ─────────────────────────────────────────────────
    ws.row_dimensions[HEADER_ROW].height = 31.5
    for col_idx, (label, _) in HEADERS.items():
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=label)
        bg  = HEADER_BG.get(col_idx)
        txt = HEADER_TXT.get(col_idx, _BLACK)
        if bg:
            cell.fill = _fill(bg)
        cell.font = Font(bold=True, size=9, name="Calibri", color=txt if bg else _BLACK)
        cell.alignment = _center(wrap=True)

    # ── Rows 6+: data ─────────────────────────────────────────────────────────
    for offset, campaign in enumerate(campaigns):
        row = DATA_START + offset
        ws.row_dimensions[row].height = 12

        J  = _L(COL_CPA);        K  = _L(COL_ER_CPA);  L  = _L(COL_ER_COST)
        M  = _L(COL_COST_TS);    N  = _L(COL_CONV)
        O  = _L(COL_COST_EUR);   P  = _L(COL_REV_EUR)
        Q  = _L(COL_MARGIN_EUR)
        AC = _L(COL_CONV_CLIENT); AD = _L(COL_REV_CLIENT)
        AE = _L(COL_VAR_CONV);   AF = _L(COL_VAR_REV)
        AG = _L(COL_VAR_CONV_PCT); AH = _L(COL_VAR_REV_PCT)

        def _w(col, value, fmt=None):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = Font(size=9, color=_GREY_DARK, name="Calibri")
            cell.alignment = _center()
            if fmt:
                cell.number_format = fmt
            elif col in NUM_FMT:
                cell.number_format = NUM_FMT[col]
            return cell

        # Static value columns
        _w(COL_MONTH,       campaign.month.month_date if campaign.month else None,
           r'[$-409]mmm\-yy')
        _w(COL_CLIENT,      campaign.client.code if campaign.client else None)
        _w(COL_COUNTRY,     campaign.country.iso_code if campaign.country else None)
        _w(COL_CATEGORIA,   campaign.category.name if campaign.category else None)
        _w(COL_CREATIVITY,  campaign.creative.name if campaign.creative else None)
        _w(COL_PLATFORM,    campaign.platform.name if campaign.platform else None)
        _w(COL_PLAT_OWNER,  campaign.plat_owner.name if campaign.plat_owner else None)
        _w(COL_CLIENT_CAMP, "YES" if campaign.client_camp else "NO")
        _w(COL_CPA,         _dec(campaign, "cpa"))
        _w(COL_ER_CPA,      _dec(campaign, "er_cpa"))
        _w(COL_ER_COST,     _dec(campaign, "er_cost"))
        _w(COL_COST_TS,     _dec(campaign, "cost_ts"))
        _w(COL_CONV,        _int(campaign, "conv"))

        # Cost/Rev/Margin written as actual values (er_cpa/er_cost null from API v2)
        _w(COL_COST_EUR,   _dec(campaign, "cost_eur"))
        _w(COL_REV_EUR,    _dec(campaign, "revenue_eur"))
        _w(COL_MARGIN_EUR, _dec(campaign, "margin_eur"))
        # ROI as formula referencing the cells above
        _w(COL_ROI, f"={Q}{row}/{O}{row}")

        # Worker columns
        _w(COL_TS_MANAGER,     _worker(campaign, "ts_manager"))
        _w(COL_TS_HEAD,        _worker(campaign, "ts_head"))
        _w(COL_CLIENT_MANAGER, _worker(campaign, "client_manager"))
        _w(COL_CLIENT_HEAD,    _worker(campaign, "client_head"))
        _w(COL_FE,             _worker(campaign, "fe"))
        _w(COL_BE,             _worker(campaign, "be"))
        _w(COL_CRIATIVO,       _worker(campaign, "creative"))
        _w(COL_QA,             _worker(campaign, "qa"))
        _w(COL_PM,             _worker(campaign, "pm"))
        _w(COL_DETALHE,        campaign.detail_type.name if campaign.detail_type else None)

        # Comparison columns
        _w(COL_CONV_CLIENT, _cmp(campaign, "client_conv"))
        _w(COL_REV_CLIENT,  _cmp(campaign, "client_rev"))

        # Var formula columns
        for col, formula in [
            (COL_VAR_CONV,    f"=+{AC}{row}-{N}{row}"),
            (COL_VAR_REV,     f"=+{AD}{row}-({N}{row})*{J}{row}"),
            (COL_VAR_CONV_PCT,f"=+{AE}{row}/{AC}{row}"),
            (COL_VAR_REV_PCT, f"=+{AF}{row}/{AD}{row}"),
            (COL_OK_CONV,     f'=+IF(AND(-0.05<{AG}{row},{AG}{row}<0.05),"ok","not ok")'),
            (COL_OK_REV,      f'=+IF(AND(-0.05<{AH}{row},{AH}{row}<0.05),"ok","not ok")'),
        ]:
            _w(col, formula)

        _w(COL_INVOICE, campaign.invoice_google)

    # Freeze panes and auto-filter
    ws.freeze_panes = f"B{DATA_START}"
    ws.auto_filter.ref = f"B{HEADER_ROW}:{_L(COL_INVOICE)}{HEADER_ROW}"


# ── Sheet 2: Pivots_2026 ──────────────────────────────────────────────────────

def _build_pivots(wb, campaigns: list) -> None:
    ws = wb.create_sheet("Pivots_2026")

    # Group campaigns by month
    from collections import defaultdict
    monthly = defaultdict(lambda: {"cost": 0.0, "rev": 0.0, "margin": 0.0})
    for c in campaigns:
        if not c.month:
            continue
        key = c.month.month_date
        cost   = _dec(c, "cost_eur")   or 0.0
        rev    = _dec(c, "revenue_eur") or 0.0
        margin = _dec(c, "margin_eur")  or 0.0
        monthly[key]["cost"]   += cost
        monthly[key]["rev"]    += rev
        monthly[key]["margin"] += margin

    # Header area
    ws.cell(row=4, column=2, value="Platform").font = Font(size=9, name="Calibri")
    ws.cell(row=4, column=3, value="Google").font   = Font(size=9, name="Calibri")

    for col, label in [(2,"Month"),(3,"Cost€"),(4,"Rev€"),(5,"Marg€"),(6,"ROI%")]:
        ws.cell(row=7, column=col, value=label).font = Font(size=9, name="Calibri")

    # Monthly rows
    total_cost = total_rev = total_margin = 0.0
    data_row = 8
    for month_date in sorted(monthly.keys()):
        d = monthly[month_date]
        roi = (d["margin"] / d["cost"]) if d["cost"] else 0.0
        total_cost   += d["cost"]
        total_rev    += d["rev"]
        total_margin += d["margin"]

        ws.cell(row=data_row, column=2, value=month_date).number_format = r'[$-409]mmm\-yy'
        ws.cell(row=data_row, column=3, value=d["cost"]).number_format   = "#,##0.00"
        ws.cell(row=data_row, column=4, value=d["rev"]).number_format    = "#,##0.00"
        ws.cell(row=data_row, column=5, value=d["margin"]).number_format = "#,##0.00"
        ws.cell(row=data_row, column=6, value=roi).number_format         = "0%"

        for col in range(2, 7):
            ws.cell(row=data_row, column=col).font = Font(size=9, name="Calibri")

        data_row += 1

    # Grand Total row
    total_roi = (total_margin / total_cost) if total_cost else 0.0
    ws.cell(row=data_row, column=2, value="Grand Total").font = Font(size=9, name="Calibri")
    ws.cell(row=data_row, column=3, value=total_cost).number_format   = "#,##0.00"
    ws.cell(row=data_row, column=4, value=total_rev).number_format    = "#,##0.00"
    ws.cell(row=data_row, column=5, value=total_margin).number_format = "#,##0.00"
    ws.cell(row=data_row, column=6, value=total_roi).number_format    = "0%"
    for col in range(2, 7):
        ws.cell(row=data_row, column=col).font = Font(size=9, name="Calibri")

    # Column widths
    for col, width in [(2,13),(3,13),(4,17),(5,16),(6,13)]:
        ws.column_dimensions[_L(col)].width = width


# ── Main entry point ──────────────────────────────────────────────────────────

def generate_monthly_excel(
    month_date: Optional[str] = None,
    client_code: Optional[str] = None,
) -> io.BytesIO:
    """
    Build the monthly closing Excel workbook matching MEC Register2026 format.

    Args:
        month_date:  "YYYY-MM" string to filter a specific month. None = all months.
        client_code: Client code to filter. None = all clients.

    Returns:
        BytesIO with .xlsx content ready for an HTTP response.
    """
    qs = (
        models.Campaign.objects
        .select_related(
            "month", "client", "country", "category", "creative",
            "platform", "plat_owner", "worker", "detail_type",
        )
        .prefetch_related("metric", "comparison")
        .order_by("month__month_date", "client__code", "country__iso_code")
    )

    if month_date:
        try:
            year, month = map(int, month_date.split("-"))
            qs = qs.filter(month__month_date__year=year, month__month_date__month=month)
        except (ValueError, AttributeError):
            pass

    if client_code:
        qs = qs.filter(client__code=client_code)

    campaigns = list(qs)

    wb = openpyxl.Workbook()
    _build_register(wb, campaigns)
    _build_pivots(wb, campaigns)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf